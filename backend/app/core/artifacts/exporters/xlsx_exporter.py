"""XLSX exporter — converts numerical artifacts into Excel spreadsheets.

Supports: financial_model, kpi_dashboard.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.artifact import ArtifactType

# ---------------------------------------------------------------------------
# Design tokens
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(
    start_color="0096FF", end_color="0096FF", fill_type="solid",
)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1A1A2E")
_SUBHEADER_FONT = Font(bold=True, size=11, color="1A1A2E")

_SUPPORTED_TYPES = {
    ArtifactType.FINANCIAL_MODEL,
    ArtifactType.KPI_DASHBOARD,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _auto_width(ws: Any) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_header_row(ws: Any, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_data_rows(
    ws: Any,
    start_row: int,
    data: list[dict[str, Any]],
    headers: list[str],
) -> int:
    """Write data rows and return next available row."""
    for i, record in enumerate(data):
        row = start_row + i
        for col_idx, header in enumerate(headers, 1):
            val = record.get(header, "")
            cell = ws.cell(row=row, column=col_idx, value=val)
            # Number formatting
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"
    return start_row + len(data)


def _write_section(
    ws: Any,
    start_row: int,
    title: str,
    data: list[dict[str, Any]],
) -> int:
    """Write a titled section with headers and data, return next row."""
    if not data:
        return start_row

    # Section title
    ws.cell(row=start_row, column=1, value=title).font = _SUBHEADER_FONT
    start_row += 1

    # Headers from first record keys
    headers = list(data[0].keys())
    _write_header_row(ws, start_row, headers)
    start_row += 1

    # Data
    next_row = _write_data_rows(ws, start_row, data, headers)
    return next_row + 1  # blank row after section


# ---------------------------------------------------------------------------
# Per-type builders
# ---------------------------------------------------------------------------


def _build_financial_model(wb: Workbook, content: dict[str, Any]) -> None:
    """Build financial model spreadsheet."""
    # Revenue Projections
    if content.get("revenue_projections"):
        ws = wb.active
        ws.title = "Revenue Projections"  # type: ignore[union-attr]
        ws.cell(row=1, column=1, value="Revenue Projections").font = _TITLE_FONT  # type: ignore[union-attr]
        data = content["revenue_projections"]
        headers = list(data[0].keys())
        _write_header_row(ws, 3, headers)
        _write_data_rows(ws, 4, data, headers)
        _auto_width(ws)

    # Cost Projections
    if content.get("cost_projections"):
        ws = wb.create_sheet("Cost Projections")
        ws.cell(row=1, column=1, value="Cost Projections").font = _TITLE_FONT
        data = content["cost_projections"]
        headers = list(data[0].keys())
        _write_header_row(ws, 3, headers)
        _write_data_rows(ws, 4, data, headers)
        _auto_width(ws)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value="Financial Summary").font = _TITLE_FONT

    row = 3
    summary_items = [
        ("Runway (months)", content.get("runway_months")),
        ("Burn Rate", content.get("burn_rate")),
    ]
    for label, value in summary_items:
        if value is not None:
            ws_sum.cell(row=row, column=1, value=label).font = _SUBHEADER_FONT
            cell = ws_sum.cell(row=row, column=2, value=value)
            if isinstance(value, float):
                cell.number_format = "#,##0.00"
            row += 1

    # Unit Economics
    ue = content.get("unit_economics", {})
    if ue:
        row += 1
        ws_sum.cell(
            row=row, column=1, value="Unit Economics",
        ).font = _SUBHEADER_FONT
        row += 1
        for key, val in ue.items():
            ws_sum.cell(row=row, column=1, value=key)
            ws_sum.cell(row=row, column=2, value=val)
            row += 1

    # Funding Scenarios
    if content.get("funding_scenarios"):
        ws_fund = wb.create_sheet("Funding Scenarios")
        ws_fund.cell(
            row=1, column=1, value="Funding Scenarios",
        ).font = _TITLE_FONT
        data = content["funding_scenarios"]
        headers = list(data[0].keys())
        _write_header_row(ws_fund, 3, headers)
        _write_data_rows(ws_fund, 4, data, headers)
        _auto_width(ws_fund)

    _auto_width(ws_sum)


def _build_kpi_dashboard(wb: Workbook, content: dict[str, Any]) -> None:
    """Build KPI dashboard spreadsheet."""
    ws = wb.active
    ws.title = "KPI Dashboard"  # type: ignore[union-attr]
    ws.cell(row=1, column=1, value="KPI Dashboard").font = _TITLE_FONT  # type: ignore[union-attr]

    metrics = content.get("metrics", [])
    if not metrics:
        ws.cell(row=3, column=1, value="No metrics defined")  # type: ignore[union-attr]
        return

    headers = ["Name", "Current", "Target", "Unit", "Trend", "Category"]
    _write_header_row(ws, 3, headers)

    for i, m in enumerate(metrics):
        row = 4 + i
        ws.cell(row=row, column=1, value=m.get("name", ""))  # type: ignore[union-attr]
        cur = m.get("current_value")
        tgt = m.get("target_value")
        ws.cell(row=row, column=2, value=cur)  # type: ignore[union-attr]
        ws.cell(row=row, column=3, value=tgt)  # type: ignore[union-attr]
        ws.cell(row=row, column=4, value=m.get("unit", ""))  # type: ignore[union-attr]
        ws.cell(row=row, column=5, value=m.get("trend", ""))  # type: ignore[union-attr]
        ws.cell(row=row, column=6, value=m.get("category", ""))  # type: ignore[union-attr]

        # Number format
        for col in (2, 3):
            cell = ws.cell(row=row, column=col)  # type: ignore[union-attr]
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

    _auto_width(ws)


_BUILDERS: dict[ArtifactType, Any] = {
    ArtifactType.FINANCIAL_MODEL: _build_financial_model,
    ArtifactType.KPI_DASHBOARD: _build_kpi_dashboard,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class XlsxExporter:
    """Export numerical artifacts to .xlsx bytes."""

    def export(
        self,
        artifact_type: ArtifactType,
        title: str,
        content: dict[str, Any],
    ) -> bytes:
        """Generate XLSX bytes from artifact content."""
        if artifact_type not in _SUPPORTED_TYPES:
            msg = (
                f"XlsxExporter does not support "
                f"{artifact_type.value}"
            )
            raise ValueError(msg)

        wb = Workbook()
        builder = _BUILDERS[artifact_type]
        builder(wb, content)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


xlsx_exporter = XlsxExporter()
